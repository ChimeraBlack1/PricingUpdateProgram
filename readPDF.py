# importing required modules 
import PyPDF2 
import csv
from openpyxl import Workbook

#create workbook object
wb = Workbook()

#grab active worksheet
ws = wb.active

#set PDF source to read from
sourceDoc = '2machineTest.pdf'

# creating a pdf file object in read binary mode
pdfFileObj = open(sourceDoc, 'rb') 

# creating a pdf reader object 
pdfReader = PyPDF2.PdfFileReader(pdfFileObj) 

# creating a page object 
pageObj = pdfReader.getPage(0) 

# extracting text from page 
extractedText = pageObj.extractText()

# split text at newline
extractedText = extractedText.split(' \n ')

# get the number of configs in the pdf
configs = len(extractedText) / 7

#create config pointer
configPointer = 0

for x in range(0, int(configs)):
  print(" ")
  print("Config " + str(x + 1) + " ...")
  itemNumber = extractedText[configPointer + 0]
  itemName = extractedText[configPointer + 1]
  itemDesc = extractedText[configPointer + 2]
  itemMAPP = extractedText[configPointer + 3]
  itemRMAP1 = extractedText[configPointer + 4]
  itemRMAP2 = extractedText[configPointer + 5]
  itemMSRP = extractedText[configPointer + 6]
  itemName = itemName.replace("\n", "")
  itemDesc = itemDesc.replace("\n","")
  itemMAPP = itemMAPP.replace("\n","")
  itemRMAP1 = itemRMAP1.replace("\n","")
  itemRMAP2 = itemRMAP2.replace("\n","")
  itemMSRP = itemMSRP.replace("\n","")
  #assign data to cells
  itemNumberCell = "A" + str(x+1)
  itemNameCell = "B" + str(x+1)
  itemDescCell = "C" + str(x+1)
  itemMAPPCell = "D" + str(x+1)
  itemRMAP1Cell = "E" + str(x+1)
  itemRMAP2Cell = "F" + str(x+1)
  itemMSRPCell = "G" + str(x+1)
  #insert data
  ws[itemNumberCell] = itemNumber
  ws[itemNameCell] = itemName
  ws[itemDescCell] = itemDesc
  ws[itemMAPPCell] = itemMAPP
  ws[itemRMAP1Cell] = itemRMAP1
  ws[itemRMAP2Cell] = itemRMAP2
  ws[itemMSRPCell] = itemMSRP
  #save
  wb.save('sample.xlsx')
  #advance pointer
  configPointer =  configPointer + 7
